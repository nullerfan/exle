import json
import openpyxl
from openpyxl import Workbook
import re

# Load JSON data
with open('/workspace/prud.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Create new workbook
wb = Workbook()
ws = wb.active

# Define headers based on the sample Excel file
headers = ['حذف شده', ' فعال  ', ' دانلودی  ', ' مجازی  ', 'شناسه', 'نام', 'SKU', 'تصویر شاخص', ' گالری تصاویر  ', 'نام لاتین', ' توضیحات سئو  ', ' تایتل اختصاصی    ', ' کلمات کلیدی    ', 'لینک', 'دسته\u200cبندی', 'برند']

# Write headers
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Simple translation function for Persian to English (basic transliteration)
def persian_to_english(text):
    """Basic Persian to English transliteration for product names"""
    # Common mappings for Persian characters
    char_map = {
        'آ': 'a', 'ا': 'a', 'ب': 'b', 'پ': 'p', 'ت': 't', 'ث': 's',
        'ج': 'j', 'چ': 'ch', 'ح': 'h', 'خ': 'kh', 'د': 'd', 'ذ': 'z',
        'ر': 'r', 'ز': 'z', 'ژ': 'zh', 'س': 's', 'ش': 'sh', 'ص': 's',
        'ض': 'z', 'ط': 't', 'ظ': 'z', 'ع': 'a', 'غ': 'gh', 'ف': 'f',
        'ق': 'gh', 'ک': 'k', 'گ': 'g', 'ل': 'l', 'م': 'm', 'ن': 'n',
        'و': 'v', 'ه': 'h', 'ی': 'y', 'ئ': 'y', 'ء': '', 'ؤ': 'o',
        ' ': ' ', '-': '-', '0': '0', '1': '1', '2': '2', '3': '3',
        '4': '4', '5': '5', '6': '6', '7': '7', '8': '8', '9': '9',
        '.': '.', '%': '%', '+': '+', '/': '/', '(': '(', ')': ')'
    }
    
    result = []
    for char in text:
        result.append(char_map.get(char, char))
    
    return ''.join(result).strip()

# Process each item from JSON
for row_idx, item in enumerate(data, 2):
    name = item.get('name', '')
    image_url = item.get('imageUrl', '')
    master_id = item.get('masterProductId', '')
    
    # Create SKU and Link with prefixes
    sku = f"SKU-{master_id}" if master_id else ''
    link = f"sku-{master_id}" if master_id else ''
    
    # Translate name to English for Latin name
    latin_name = persian_to_english(name)
    
    # Fill only the specified columns
    # Column indices (1-based): 
    # 5=شناسه (leave empty), 6=نام, 7=SKU, 8=تصویر شاخص, 9=گالری تصاویر (empty)
    # 10=نام لاتین, 11=توضیحات سئو (empty), 12=تایتل اختصاصی (empty), 14=لینک
    
    ws.cell(row=row_idx, column=6, value=name)  # نام
    ws.cell(row=row_idx, column=7, value=sku)   # SKU
    ws.cell(row=row_idx, column=8, value=image_url)  # تصویر شاخص
    # Column 9 (گالری تصاویر) - leave empty
    ws.cell(row=row_idx, column=10, value=latin_name)  # نام لاتین
    # Column 11 (توضیحات سئو) - leave empty
    # Column 12 (تایتل اختصاصی) - leave empty
    ws.cell(row=row_idx, column=14, value=link)  # لینک
    
    # Leave other columns empty (شناسه, دسته‌بندی, برند, etc.)

# Save the workbook
wb.save('/workspace/prud.xlsx')
print(f"Successfully created prud.xlsx with {len(data)} products")
