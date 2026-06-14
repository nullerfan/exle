import json
import openpyxl
from openpyxl import Workbook
import re

# خواندن فایل JSON
with open('/workspace/prud.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# خواندن ساختار از فایل Products.xlsx
wb_template = openpyxl.load_workbook('/workspace/Products.xlsx')
ws_template = wb_template.active
columns = [cell.value for cell in ws_template[1]]

# ایجاد فایل جدید
wb = Workbook()
ws = wb.active

# نوشتن هدرها
for col_idx, col_name in enumerate(columns, 1):
    ws.cell(row=1, column=col_idx, value=col_name)

def persian_to_latin(persian_text):
    """تبدیل متن فارسی به لاتین برای نام لاتین"""
    if not persian_text:
        return ""
    
    # حذف کاراکترهای غیر مجاز و فاصله‌های اضافی
    text = persian_text.strip()
    
    # نگاشت حروف فارسی به لاتین
    char_map = {
        'آ': 'a', 'ا': 'a', 'أ': 'a', 'إ': 'a',
        'ب': 'b', 'پ': 'p', 'ت': 't', 'ث': 's',
        'ج': 'j', 'چ': 'ch', 'ح': 'h', 'خ': 'kh',
        'د': 'd', 'ذ': 'z', 'ر': 'r', 'ز': 'z',
        'ژ': 'zh', 'س': 's', 'ش': 'sh', 'ص': 's',
        'ض': 'z', 'ط': 't', 'ظ': 'z', 'ع': 'a',
        'غ': 'gh', 'ف': 'f', 'ق': 'gh', 'ک': 'k',
        'گ': 'g', 'ل': 'l', 'م': 'm', 'ن': 'n',
        'و': 'v', 'ه': 'h', 'ی': 'y', 'ي': 'y',
        'ئ': 'y', 'ء': '', 'ة': 'h',
        ' ': '-', '0': '0', '1': '1', '2': '2', '3': '3',
        '4': '4', '5': '5', '6': '6', '7': '7', '8': '8', '9': '9',
        '.': '-', '(': '', ')': '', '/': '-', '+': '-', '*': '-'
    }
    
    result = []
    for char in text:
        if char in char_map:
            result.append(char_map[char])
        elif char.isalnum():
            result.append(char.lower())
    
    # حذف خط تیره‌های متوالی و انتهایی
    latin = ''.join(result)
    latin = re.sub(r'-+', '-', latin)
    latin = latin.strip('-')
    
    return latin.lower()[:100] if latin else "product"

def generate_seo_description(name, master_product_id):
    """تولید توضیحات سئو بر اساس نام محصول"""
    if not name:
        return ""
    
    # حذف اعداد و کاراکترهای خاص از نام برای توضیحات
    clean_name = re.sub(r'\d+\s*(لیتری|گرمی|کیلوگرمی|عدد|بسته)', '', name).strip()
    
    description = f"{clean_name} با کیفیت عالی. خرید آنلاین {clean_name} با بهترین قیمت."
    return description[:255] if len(description) > 255 else description

# پردازش هر رکورد
for row_idx, item in enumerate(data, 2):
    master_id = item.get('masterProductId', '')
    name = item.get('name', '')
    image_url = item.get('imageUrl', '')
    
    # تولید نام لاتین
    latin_name = persian_to_latin(name)
    
    # تولید توضیحات سئو
    seo_desc = generate_seo_description(name, master_id)
    
    # SKU با پیشوند sku-
    sku = f"sku-{master_id}"
    
    row_data = {
        'حذف شده': 0,
        ' فعال  ': 1,
        ' دانلودی  ': 0,
        ' مجازی  ': 0,
        'شناسه': master_id,
        'نام': name,
        'SKU': sku,
        'تصویر شاخص': image_url,
        ' گالری تصاویر  ': '',
        'نام لاتین': latin_name,
        ' توضیحات سئو  ': seo_desc,
        ' تایتل اختصاصی    ': '',
        ' کلمات کلیدی    ': '',
        'لینک': '',
        'دسته\u200cبندی': '',
        'برند': ''
    }
    
    for col_idx, col_name in enumerate(columns, 1):
        col_name_clean = col_name.strip() if col_name else ''
        # پیدا کردن کلید مناسب در row_data
        value = None
        for key, val in row_data.items():
            if key.strip() == col_name_clean:
                value = val
                break
        ws.cell(row=row_idx, column=col_idx, value=value)

# ذخیره فایل
wb.save('/workspace/prud.xlsx')
print(f"فایل prud.xlsx با موفقیت ایجاد شد. تعداد رکوردها: {len(data)}")
