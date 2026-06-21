import json
import openpyxl
from openpyxl import Workbook

# Load JSON data
with open('/workspace/prud.json', 'r', encoding='utf-8') as f:
    products = json.load(f)

# Create a new workbook
wb = Workbook()
ws = wb.active

# Define headers based on the sample Excel file
headers = [
    'حذف شده',
    ' فعال  ',
    ' دانلودی  ',
    ' مجازی  ',
    'شناسه',
    'نام',
    'SKU',
    'تصویر شاخص',
    ' گالری تصاویر  ',
    'نام لاتین',
    ' توضیحات سئو  ',
    ' تایتل اختصاصی    ',
    ' کلمات کلیدی    ',
    'لینک',
    'دسته\u200cبندی',
    'برند'
]

# Write headers to first row
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Simple function to translate Persian to English (basic transliteration for product names)
def simple_translate_to_english(persian_text):
    """
    A simple approach to convert Persian product names to English.
    Since we don't have access to a full translation API, we'll create a basic mapping
    or use a simplified approach.
    """
    # For this task, we'll use a basic character mapping
    persian_to_english = {
        'آ': 'A', 'ا': 'a', 'ب': 'b', 'پ': 'p', 'ت': 't', 'ث': 's',
        'ج': 'j', 'چ': 'ch', 'ح': 'h', 'خ': 'kh', 'د': 'd', 'ذ': 'z',
        'ر': 'r', 'ز': 'z', 'ژ': 'zh', 'س': 's', 'ش': 'sh', 'ص': 's',
        'ض': 'z', 'ط': 't', 'ظ': 'z', 'ع': 'a', 'غ': 'gh', 'ف': 'f',
        'ق': 'gh', 'ک': 'k', 'گ': 'g', 'ل': 'l', 'م': 'm', 'ن': 'n',
        'و': 'v', 'ه': 'h', 'ی': 'y', 'ئ': 'y', 'ء': '', ' ': ' ',
        '۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4',
        '۵': '5', '۶': '6', '۷': '7', '۸': '8', '۹': '9',
        '0': '0', '1': '1', '2': '2', '3': '3', '4': '4',
        '5': '5', '6': '6', '7': '7', '8': '8', '9': '9',
        '.': '.', '(': '', ')': '', '-': '-', '%': '%', '+': '+'
    }
    
    result = []
    for char in persian_text:
        result.append(persian_to_english.get(char, char))
    
    return ''.join(result).strip()

# Process each product and fill only specified columns
for row_idx, product in enumerate(products, 2):
    # Get values from JSON
    name = product.get('name', '')
    image_url = product.get('imageUrl', '')
    master_product_id = product.get('masterProductId', '')
    
    # Create SKU with prefix "SKU-"
    sku = f"SKU-{master_product_id}"
    
    # Create link with prefix "sku-"
    link = f"sku-{master_product_id}"
    
    # Translate name to English for Latin name
    latin_name = simple_translate_to_english(name)
    
    # Fill only the specified columns:
    # Column F (6): نام (Name)
    ws.cell(row=row_idx, column=6, value=name)
    
    # Column G (7): SKU
    ws.cell(row=row_idx, column=7, value=sku)
    
    # Column H (8): تصویر شاخص (Main Image)
    ws.cell(row=row_idx, column=8, value=image_url)
    
    # Column I (9): گالری تصاویر (Image Gallery) - leaving empty as not specified
    # Column J (10): نام لاتین (Latin Name)
    ws.cell(row=row_idx, column=10, value=latin_name)
    
    # Columns K, L (11, 12): توضیحات سئو, تایتل اختصاصی - leaving empty as not specified
    # Column N (14): لینک (Link)
    ws.cell(row=row_idx, column=14, value=link)
    
    # Other columns (شناسه, دسته‌بندی, برند, etc.) are left empty as requested

# Save the workbook
wb.save('/workspace/prud.xlsx')
print(f"Successfully created prud.xlsx with {len(products)} products")
